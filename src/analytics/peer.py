"""Peer group percentile rankings, radar charts, and comparison workbooks."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import re
import sqlite3
from typing import Any

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
from openpyxl.styles import Font, PatternFill
import pandas as pd

from src.screener.engine import DB_PATH, prepare_screener_universe, load_financial_ratios


ROOT = Path(__file__).resolve().parents[2]
PEER_GROUPS_PATH = ROOT / "data" / "raw" / "peer_groups.xlsx"
PEER_REPORT_PATH = ROOT / "output" / "peer_comparison.xlsx"
RADAR_DIR = ROOT / "reports" / "radar_charts"


@dataclass(frozen=True)
class PeerMetric:
    metric: str
    column: str
    higher_is_better: bool = True


PEER_RANK_METRICS = (
    PeerMetric("ROE", "return_on_equity_pct"),
    PeerMetric("ROCE", "return_on_capital_employed_pct"),
    PeerMetric("Net Profit Margin", "net_profit_margin_pct"),
    PeerMetric("D/E", "debt_to_equity", higher_is_better=False),
    PeerMetric("FCF", "free_cash_flow_cr"),
    PeerMetric("PAT CAGR 5yr", "pat_cagr_5yr"),
    PeerMetric("Revenue CAGR 5yr", "revenue_cagr_5yr"),
    PeerMetric("EPS CAGR 5yr", "eps_cagr_5yr"),
    PeerMetric("Interest Coverage", "interest_coverage"),
    PeerMetric("Asset Turnover", "asset_turnover"),
)

RADAR_METRICS = (
    PeerMetric("ROE", "return_on_equity_pct"),
    PeerMetric("ROCE", "return_on_capital_employed_pct"),
    PeerMetric("NPM", "net_profit_margin_pct"),
    PeerMetric("D/E", "debt_to_equity", higher_is_better=False),
    PeerMetric("FCF score", "free_cash_flow_cr"),
    PeerMetric("PAT CAGR 5yr", "pat_cagr_5yr"),
    PeerMetric("Revenue CAGR 5yr", "revenue_cagr_5yr"),
    PeerMetric("Composite Score", "composite_quality_score"),
)

REPORT_METRIC_COLUMNS = (
    "composite_quality_score",
    "return_on_equity_pct",
    "return_on_capital_employed_pct",
    "net_profit_margin_pct",
    "debt_to_equity",
    "free_cash_flow_cr",
    "pat_cagr_5yr",
    "revenue_cagr_5yr",
    "eps_cagr_5yr",
    "interest_coverage",
    "asset_turnover",
    "pe_ratio",
    "pb_ratio",
    "dividend_yield",
    "dividend_payout_ratio_pct",
    "cash_from_operations_cr",
    "sales",
    "net_profit",
    "market_cap_crore",
    "fcf_cagr_5yr",
)


def load_peer_groups(path: str | Path = PEER_GROUPS_PATH) -> pd.DataFrame:
    """Load analyst-maintained peer group assignments."""
    peers = pd.read_excel(path)
    expected = {"peer_group_name", "company_id", "is_benchmark"}
    missing = expected.difference(peers.columns)
    if missing:
        raise ValueError(f"peer_groups.xlsx missing columns: {', '.join(sorted(missing))}")
    peers = peers.loc[:, ["peer_group_name", "company_id", "is_benchmark"]].copy()
    peers["company_id"] = peers["company_id"].astype("string").str.strip()
    peers["company_id"] = peers["company_id"].str.replace("&", "", regex=False)
    peers["peer_group_name"] = peers["peer_group_name"].astype("string").str.strip()
    peers["is_benchmark"] = peers["is_benchmark"].astype(bool)
    return peers


def build_peer_universe(
    db_path: str | Path = DB_PATH,
    peer_groups_path: str | Path = PEER_GROUPS_PATH,
) -> pd.DataFrame:
    """Return latest annual company metrics joined to peer group assignments."""
    universe = prepare_screener_universe(load_financial_ratios(db_path))
    peers = load_peer_groups(peer_groups_path)
    return peers.merge(universe, on="company_id", how="left", validate="one_to_one")


def compute_peer_percentiles(peer_universe: pd.DataFrame) -> pd.DataFrame:
    """Compute 0-1 percentile ranks for configured metrics within each peer group."""
    rows: list[dict[str, Any]] = []
    for peer_group_name, group in peer_universe.groupby("peer_group_name", sort=True):
        for metric in PEER_RANK_METRICS:
            if metric.column in group.columns:
                values = pd.to_numeric(group[metric.column], errors="coerce")
            else:
                values = pd.Series(pd.NA, index=group.index, dtype="Float64")
            valid = values.dropna()
            percentiles = pd.Series(pd.NA, index=group.index, dtype="Float64")
            if len(valid) <= 1:
                percentiles.loc[valid.index] = 1.0
            else:
                percentiles.loc[valid.index] = (valid.rank(method="average", ascending=True) - 1) / (len(valid) - 1)
            if not metric.higher_is_better:
                percentiles.loc[valid.index] = 1 - percentiles.loc[valid.index]

            for index, source_row in group.iterrows():
                percentile = percentiles.loc[index]
                source_row = group.loc[index]
                rows.append(
                    {
                        "company_id": source_row["company_id"],
                        "peer_group_name": peer_group_name,
                        "metric": metric.metric,
                        "value": values.loc[index],
                        "percentile_rank": round(float(percentile), 6) if pd.notna(percentile) else pd.NA,
                        "year": source_row.get("year"),
                    }
                )
    return pd.DataFrame(rows)


def populate_peer_percentiles(
    db_path: str | Path = DB_PATH,
    peer_groups_path: str | Path = PEER_GROUPS_PATH,
) -> int:
    """Create and populate peer_percentiles in SQLite."""
    peer_universe = build_peer_universe(db_path, peer_groups_path)
    percentiles = compute_peer_percentiles(peer_universe)
    with sqlite3.connect(db_path) as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS peer_percentiles (
                company_id TEXT,
                peer_group_name TEXT,
                metric TEXT,
                value REAL,
                percentile_rank REAL,
                year TEXT
            )
            """
        )
        conn.execute("DELETE FROM peer_percentiles")
        percentiles.to_sql("peer_percentiles", conn, if_exists="append", index=False)
        conn.commit()
    return len(percentiles)


def peer_group_message(company_id: str, peer_groups_path: str | Path = PEER_GROUPS_PATH) -> str:
    """Return a friendly message for companies outside configured peer groups."""
    peers = load_peer_groups(peer_groups_path)
    if company_id not in set(peers["company_id"]):
        return "No peer group assigned"
    group_name = peers.loc[peers["company_id"].eq(company_id), "peer_group_name"].iloc[0]
    return f"Peer group assigned: {group_name}"


def export_peer_comparison(
    output_path: str | Path = PEER_REPORT_PATH,
    db_path: str | Path = DB_PATH,
    peer_groups_path: str | Path = PEER_GROUPS_PATH,
) -> dict[str, int]:
    """Generate the peer comparison workbook with one sheet per peer group."""
    peer_universe = build_peer_universe(db_path, peer_groups_path)
    percentiles = compute_peer_percentiles(peer_universe)
    path = Path(output_path)
    path.parent.mkdir(exist_ok=True)

    counts: dict[str, int] = {}
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for group_name, group in peer_universe.groupby("peer_group_name", sort=True):
            report = _build_report_sheet(group, percentiles)
            counts[str(group_name)] = len(group)
            report.to_excel(writer, sheet_name=_sheet_name(str(group_name)), index=False)
            _format_peer_sheet(writer.book[_sheet_name(str(group_name))], report)
    return counts


def generate_radar_charts(
    output_dir: str | Path = RADAR_DIR,
    db_path: str | Path = DB_PATH,
    peer_groups_path: str | Path = PEER_GROUPS_PATH,
) -> int:
    """Generate peer overlay radar charts and standalone charts for unassigned companies."""
    output = Path(output_dir)
    output.mkdir(parents=True, exist_ok=True)
    for existing in output.glob("*_radar.png"):
        existing.unlink()
    universe = prepare_screener_universe(load_financial_ratios(db_path))
    peers = load_peer_groups(peer_groups_path)
    grouped = peers.merge(universe, on="company_id", how="left", validate="one_to_one")
    assigned_ids = set(grouped["company_id"])
    chart_count = 0

    for group_name, group in grouped.groupby("peer_group_name", sort=True):
        scores = _metric_scores(group, RADAR_METRICS)
        peer_average = scores.mean(axis=0, skipna=True).fillna(0)
        for idx, row in group.iterrows():
            company_scores = scores.loc[idx].fillna(0)
            _save_radar_chart(
                output / f"{row['company_id']}_radar.png",
                str(row["company_id"]),
                str(group_name),
                [metric.metric for metric in RADAR_METRICS],
                company_scores.tolist(),
                peer_average.tolist(),
            )
            chart_count += 1

    nifty_average = pd.to_numeric(universe["composite_quality_score"], errors="coerce").mean()
    unassigned = universe[~universe["company_id"].isin(assigned_ids)]
    for _, row in unassigned.iterrows():
        _save_standalone_chart(
            output / f"{row['company_id']}_radar.png",
            str(row["company_id"]),
            float(row["composite_quality_score"]),
            float(nifty_average),
        )
        chart_count += 1
    return chart_count


def run_peer_analytics(
    db_path: str | Path = DB_PATH,
    peer_groups_path: str | Path = PEER_GROUPS_PATH,
    report_path: str | Path = PEER_REPORT_PATH,
    radar_dir: str | Path = RADAR_DIR,
) -> dict[str, Any]:
    """Run all Day 18-20 peer analytics deliverables."""
    percentile_rows = populate_peer_percentiles(db_path, peer_groups_path)
    sheet_counts = export_peer_comparison(report_path, db_path, peer_groups_path)
    chart_count = generate_radar_charts(radar_dir, db_path, peer_groups_path)
    return {
        "peer_percentile_rows": percentile_rows,
        "peer_group_sheets": len(sheet_counts),
        "radar_charts": chart_count,
    }


def _build_report_sheet(group: pd.DataFrame, percentiles: pd.DataFrame) -> pd.DataFrame:
    base_columns = ["company_id", "company_name", "is_benchmark", *REPORT_METRIC_COLUMNS]
    report = group.reindex(columns=base_columns).copy()
    for metric in PEER_RANK_METRICS:
        rank_column = f"{metric.metric} Percentile"
        metric_ranks = percentiles[
            percentiles["peer_group_name"].eq(group["peer_group_name"].iloc[0])
            & percentiles["metric"].eq(metric.metric)
        ].set_index("company_id")["percentile_rank"]
        report[rank_column] = report["company_id"].map(metric_ranks)

    medians = {column: pd.NA for column in report.columns}
    medians["company_id"] = "Median"
    medians["company_name"] = "Peer group median"
    for column in REPORT_METRIC_COLUMNS:
        medians[column] = pd.to_numeric(report[column], errors="coerce").median()
    for metric in PEER_RANK_METRICS:
        medians[f"{metric.metric} Percentile"] = pd.to_numeric(report[f"{metric.metric} Percentile"], errors="coerce").median()
    return pd.concat([report, pd.DataFrame([medians])], ignore_index=True)


def _metric_scores(group: pd.DataFrame, metrics: tuple[PeerMetric, ...]) -> pd.DataFrame:
    scores = pd.DataFrame(index=group.index)
    for metric in metrics:
        values = pd.to_numeric(group[metric.column], errors="coerce")
        valid = values.dropna()
        if len(valid) <= 1:
            percentile = pd.Series(100.0, index=valid.index)
        else:
            percentile = (valid.rank(method="average", ascending=True) - 1) / (len(valid) - 1) * 100
        if not metric.higher_is_better:
            percentile = 100 - percentile
        scores[metric.metric] = percentile
    return scores


def _save_radar_chart(
    path: Path,
    company_id: str,
    peer_group_name: str,
    labels: list[str],
    company_values: list[float],
    peer_values: list[float],
) -> None:
    angles = np.linspace(0, 2 * np.pi, len(labels), endpoint=False).tolist()
    angles += angles[:1]
    company_plot = company_values + company_values[:1]
    peer_plot = peer_values + peer_values[:1]

    fig, ax = plt.subplots(figsize=(8, 8), subplot_kw={"polar": True})
    ax.plot(angles, company_plot, color="#2563eb", linewidth=2)
    ax.fill(angles, company_plot, color="#2563eb", alpha=0.22)
    ax.plot(angles, peer_plot, color="#111827", linewidth=2, linestyle="--")
    ax.set_ylim(0, 100)
    ax.set_xticks(angles[:-1])
    ax.set_xticklabels(labels, fontsize=9)
    ax.set_yticks([25, 50, 75, 100])
    ax.set_yticklabels(["25", "50", "75", "100"], fontsize=8)
    ax.set_title(f"{company_id} vs {peer_group_name}", fontsize=13, pad=22)
    ax.legend(["Company", "Peer average"], loc="lower center", bbox_to_anchor=(0.5, -0.14), ncol=2)
    fig.tight_layout()
    fig.savefig(path, dpi=150, bbox_inches="tight")
    plt.close(fig)


def _save_standalone_chart(path: Path, company_id: str, company_score: float, nifty_average: float) -> None:
    fig, ax = plt.subplots(figsize=(7, 4.2))
    ax.bar(["Company", "Nifty 100 avg"], [company_score, nifty_average], color=["#2563eb", "#f59e0b"])
    ax.set_ylim(0, 100)
    ax.set_ylabel("Composite score", fontsize=10)
    ax.set_title(f"{company_id} standalone reference", fontsize=13)
    ax.grid(axis="y", alpha=0.25)
    fig.tight_layout()
    fig.savefig(path, dpi=150, bbox_inches="tight")
    plt.close(fig)


def _format_peer_sheet(worksheet: Any, report: pd.DataFrame) -> None:
    green_fill = PatternFill(fill_type="solid", fgColor="C6EFCE")
    yellow_fill = PatternFill(fill_type="solid", fgColor="FFEB9C")
    red_fill = PatternFill(fill_type="solid", fgColor="FFC7CE")
    amber_fill = PatternFill(fill_type="solid", fgColor="FCD34D")
    median_fill = PatternFill(fill_type="solid", fgColor="E5E7EB")
    worksheet.freeze_panes = "A2"

    headers = {cell.value: cell.column for cell in worksheet[1]}
    percentile_columns = [column for column in report.columns if column.endswith(" Percentile")]
    benchmark_column = headers.get("is_benchmark")
    for row_idx in range(2, worksheet.max_row + 1):
        is_median = worksheet.cell(row=row_idx, column=1).value == "Median"
        is_benchmark = benchmark_column and worksheet.cell(row=row_idx, column=benchmark_column).value is True
        if is_median or is_benchmark:
            row_fill = median_fill if is_median else amber_fill
            for col_idx in range(1, worksheet.max_column + 1):
                worksheet.cell(row=row_idx, column=col_idx).fill = row_fill
                if is_median:
                    worksheet.cell(row=row_idx, column=col_idx).font = Font(bold=True)
            continue

        for column in percentile_columns:
            col_idx = headers[column]
            cell = worksheet.cell(row=row_idx, column=col_idx)
            value = pd.to_numeric(cell.value, errors="coerce")
            if pd.isna(value):
                continue
            if value >= 0.75:
                cell.fill = green_fill
            elif value <= 0.25:
                cell.fill = red_fill
            else:
                cell.fill = yellow_fill

    for cell in worksheet[1]:
        cell.font = Font(bold=True)
    for column_cells in worksheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells[:100])
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 10), 24)


def _sheet_name(group_name: str) -> str:
    cleaned = re.sub(r"[\[\]:*?/\\]", " ", group_name).strip()
    return cleaned[:31]


if __name__ == "__main__":
    print(run_peer_analytics())
