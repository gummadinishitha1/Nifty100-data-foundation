"""Threshold-based financial screener engine."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import sqlite3
from typing import Any

from openpyxl.styles import PatternFill
import pandas as pd
import yaml

from src.analytics.cagr import calculate_cagr, fiscal_year_sort_key


ROOT = Path(__file__).resolve().parents[2]
DB_PATH = ROOT / "data" / "db" / "nifty100.db"
CONFIG_PATH = ROOT / "config" / "screener_config.yaml"
OUTPUT_PATH = ROOT / "output" / "screener_output.xlsx"


@dataclass(frozen=True)
class FilterSpec:
    """Column aliases and comparison direction for a screener threshold."""

    aliases: tuple[str, ...]
    direction: str


FILTER_SPECS: dict[str, FilterSpec] = {
    "roe_min": FilterSpec(("return_on_equity_pct", "roe"), "min"),
    "de_max": FilterSpec(("debt_to_equity",), "max"),
    "de_eq": FilterSpec(("debt_to_equity",), "eq"),
    "fcf_min": FilterSpec(("free_cash_flow_cr",), "min"),
    "latest_fcf_positive": FilterSpec(("fcf_positive_flag",), "min"),
    "revenue_cagr_5yr_min": FilterSpec(("revenue_cagr_5yr",), "min"),
    "revenue_cagr_3yr_min": FilterSpec(("revenue_cagr_3yr",), "min"),
    "pat_cagr_5yr_min": FilterSpec(("pat_cagr_5yr",), "min"),
    "opm_min": FilterSpec(("operating_profit_margin_pct", "opm"), "min"),
    "pe_max": FilterSpec(("pe_ratio",), "max"),
    "pb_max": FilterSpec(("pb_ratio",), "max"),
    "dividend_yield_min": FilterSpec(("dividend_yield",), "min"),
    "dividend_payout_max": FilterSpec(("dividend_payout_ratio_pct",), "max"),
    "icr_min": FilterSpec(("interest_coverage",), "min"),
    "market_cap_min": FilterSpec(("market_cap_crore",), "min"),
    "net_profit_min": FilterSpec(("net_profit",), "min"),
    "eps_cagr_min": FilterSpec(("eps_cagr_5yr", "eps_growth"), "min"),
    "asset_turnover_min": FilterSpec(("asset_turnover",), "min"),
    "sales_min": FilterSpec(("sales",), "min"),
    "revenue_min": FilterSpec(("sales",), "min"),
    "de_declining_yoy": FilterSpec(("de_declining_yoy_flag",), "min"),
}

SCORE_COMPONENTS = (
    ("return_on_equity_pct", 15.0, 1),
    ("return_on_capital_employed_pct", 10.0, 1),
    ("net_profit_margin_pct", 10.0, 1),
    ("fcf_cagr_5yr", 15.0, 1),
    ("cfo_pat_ratio", 10.0, 1),
    ("fcf_positive_flag", 5.0, 1),
    ("revenue_cagr_5yr", 10.0, 1),
    ("pat_cagr_5yr", 10.0, 1),
    ("debt_to_equity", 10.0, -1),
    ("interest_coverage", 5.0, 1),
)

OUTPUT_COLUMNS = (
    "company_id",
    "company_name",
    "broad_sector",
    "year",
    "composite_quality_score",
    "return_on_equity_pct",
    "return_on_capital_employed_pct",
    "net_profit_margin_pct",
    "fcf_cagr_5yr",
    "cfo_pat_ratio",
    "fcf_positive_flag",
    "revenue_cagr_5yr",
    "pat_cagr_5yr",
    "debt_to_equity",
    "interest_coverage",
    "pe_ratio",
    "pb_ratio",
    "dividend_yield",
    "dividend_payout_ratio_pct",
    "free_cash_flow_cr",
    "cash_from_operations_cr",
    "sales",
    "revenue_cagr_3yr",
    "de_declining_yoy_flag",
)


def load_screener_config(config_path: str | Path = CONFIG_PATH) -> dict[str, Any]:
    """Load screener presets from YAML."""
    path = Path(config_path)
    if not path.exists():
        raise FileNotFoundError(f"Screener config not found: {path}")

    with path.open("r", encoding="utf-8") as file:
        config = yaml.safe_load(file) or {}

    if "presets" not in config or not isinstance(config["presets"], dict):
        raise ValueError("screener_config.yaml must contain a presets mapping")
    return config


def load_financial_ratios(db_path: str | Path = DB_PATH) -> pd.DataFrame:
    """Load latest financial ratios enriched with filter support columns."""
    path = Path(db_path)
    if not path.exists():
        raise FileNotFoundError(f"Database not found: {path}")

    with sqlite3.connect(path) as conn:
        return pd.read_sql_query(
            """
            SELECT
                fr.*,
                c.company_name,
                s.broad_sector,
                mc.market_cap_crore,
                mc.pe_ratio AS market_pe_ratio,
                mc.pb_ratio AS market_pb_ratio,
                mc.dividend_yield_pct AS market_dividend_yield,
                p.sales,
                p.net_profit
            FROM financial_ratios fr
            LEFT JOIN companies c
                ON c.id = fr.company_id
            LEFT JOIN sectors s
                ON s.company_id = fr.company_id
            LEFT JOIN market_cap mc
                ON mc.company_id = fr.company_id
               AND CAST(fr.year AS TEXT) LIKE '%' || CAST(mc.year AS TEXT) || '%'
            LEFT JOIN profitandloss p
                ON p.company_id = fr.company_id
               AND CAST(p.year AS TEXT) = CAST(fr.year AS TEXT)
            """,
            conn,
        )


def get_preset_thresholds(
    preset: str | None = None,
    config_path: str | Path = CONFIG_PATH,
) -> dict[str, float]:
    """Return thresholds for a configured preset."""
    config = load_screener_config(config_path)
    preset_name = preset or config.get("default_preset")
    presets = config["presets"]

    if preset_name not in presets:
        raise KeyError(f"Unknown screener preset: {preset_name}")

    thresholds = presets[preset_name].get("thresholds", {})
    return _normalise_thresholds(thresholds)


def apply_filters(
    financial_ratios: pd.DataFrame,
    thresholds: dict[str, float | int | None],
) -> pd.DataFrame:
    """Apply threshold filters to a financial ratios DataFrame."""
    result = prepare_screener_universe(financial_ratios)
    active_thresholds = _normalise_thresholds(thresholds)

    mask = pd.Series(True, index=result.index)
    for filter_name, threshold in active_thresholds.items():
        spec = _get_filter_spec(filter_name)
        column = _resolve_column(result, spec.aliases)
        values = pd.to_numeric(result[column], errors="coerce")

        if spec.direction == "min":
            filter_mask = values >= threshold
        elif spec.direction == "max":
            filter_mask = values <= threshold
        else:
            filter_mask = values == threshold

        if filter_name == "de_max":
            filter_mask = _apply_financials_de_skip(result, filter_mask)
        elif filter_name == "icr_min":
            filter_mask = _apply_debt_free_icr_pass(result, filter_mask)

        mask &= filter_mask.fillna(False)

    filtered = result.loc[mask].copy()
    return sort_screener_output(filtered)


def run_screener(
    preset: str | None = None,
    custom_thresholds: dict[str, float | int | None] | None = None,
    financial_ratios: pd.DataFrame | None = None,
    config_path: str | Path = CONFIG_PATH,
    db_path: str | Path = DB_PATH,
) -> pd.DataFrame:
    """Run a preset screener with optional custom threshold overrides."""
    thresholds = get_preset_thresholds(preset, config_path)
    if custom_thresholds:
        thresholds.update(_normalise_thresholds(custom_thresholds))

    ratios = financial_ratios if financial_ratios is not None else load_financial_ratios(db_path)
    return apply_filters(ratios, thresholds)


def prepare_screener_universe(df: pd.DataFrame) -> pd.DataFrame:
    """Return latest annual company rows with derived screener metrics and scores."""
    result = df.copy()
    result = _add_derived_metrics(result)
    result = _latest_annual_rows(result)
    result = _limit_to_sector_universe(result)
    return add_composite_quality_score(result)


def add_composite_quality_score(df: pd.DataFrame) -> pd.DataFrame:
    """Compute the sector-relative weighted composite quality score."""
    result = _ensure_score_aliases(df.copy())
    for column, weight, direction in SCORE_COMPONENTS:
        if column not in result.columns:
            result[column] = pd.NA
        result[f"{column}__score"] = _sector_relative_score(result, column, direction)

    total = pd.Series(0.0, index=result.index)
    available_weight = pd.Series(0.0, index=result.index)
    for column, weight, _ in SCORE_COMPONENTS:
        score = result[f"{column}__score"]
        present = score.notna()
        total = total.add(score.fillna(0) * weight, fill_value=0)
        available_weight = available_weight.add(present.astype(float) * weight, fill_value=0)

    result["composite_quality_score"] = (total / available_weight).where(available_weight > 0)
    result["composite_quality_score"] = result["composite_quality_score"].round(2)
    helper_columns = [f"{column}__score" for column, _, _ in SCORE_COMPONENTS]
    result = result.drop(columns=helper_columns)
    return result


def _ensure_score_aliases(df: pd.DataFrame) -> pd.DataFrame:
    aliases = {
        "return_on_equity_pct": "roe",
        "return_on_capital_employed_pct": "roce",
        "net_profit_margin_pct": "npm",
    }
    for target, source in aliases.items():
        if target not in df.columns and source in df.columns:
            df[target] = df[source]
    if "cfo_pat_ratio" not in df.columns and "cfo_quality_5yr" in df.columns:
        df["cfo_pat_ratio"] = df["cfo_quality_5yr"]
    return df


def export_screener_output(
    output_path: str | Path = OUTPUT_PATH,
    config_path: str | Path = CONFIG_PATH,
    db_path: str | Path = DB_PATH,
) -> dict[str, int]:
    """Generate an Excel workbook with one colour-coded sheet per preset."""
    config = load_screener_config(config_path)
    ratios = load_financial_ratios(db_path)
    path = Path(output_path)
    path.parent.mkdir(exist_ok=True)

    counts: dict[str, int] = {}
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for preset_name in config["presets"]:
            thresholds = get_preset_thresholds(preset_name, config_path)
            result = apply_filters(ratios, thresholds)
            counts[preset_name] = len(result)
            output = result.reindex(columns=[column for column in OUTPUT_COLUMNS if column in result.columns])
            output.to_excel(writer, sheet_name=_sheet_name(preset_name), index=False)
            _colour_code_sheet(writer.book[_sheet_name(preset_name)], output, thresholds)
            _format_sheet(writer.book[_sheet_name(preset_name)])
    return counts


def _add_derived_metrics(df: pd.DataFrame) -> pd.DataFrame:
    result = df.copy()
    for target, source in {
        "pe_ratio": "market_pe_ratio",
        "pb_ratio": "market_pb_ratio",
        "dividend_yield": "market_dividend_yield",
    }.items():
        if target in result.columns and source in result.columns:
            result[target] = result[target].combine_first(result[source])
    if "cash_from_operations_cr" in result.columns and "net_profit" in result.columns:
        cfo = pd.to_numeric(result["cash_from_operations_cr"], errors="coerce")
        pat = pd.to_numeric(result["net_profit"], errors="coerce")
        result["cfo_pat_ratio"] = (cfo / pat).where(pat != 0)
    if "free_cash_flow_cr" in result.columns:
        fcf = pd.to_numeric(result["free_cash_flow_cr"], errors="coerce")
        result["fcf_positive_flag"] = (fcf > 0).astype(int)
    else:
        result["fcf_positive_flag"] = 0
    result["de_declining_yoy_flag"] = _de_declining_yoy(result)
    result["fcf_cagr_5yr"] = _fcf_cagr(result)
    return result


def _latest_annual_rows(df: pd.DataFrame) -> pd.DataFrame:
    if "year" not in df.columns:
        return df.copy()

    result = df.copy()
    annual = result[~result["year"].astype("string").str.strip().str.casefold().eq("ttm")].copy()
    if annual.empty or "company_id" not in annual.columns:
        return result

    annual["_year_sort_key"] = annual["year"].map(fiscal_year_sort_key)
    latest_index = annual.sort_values("_year_sort_key").groupby("company_id", sort=False).tail(1).index
    return annual.loc[latest_index].drop(columns="_year_sort_key").reset_index(drop=True)


def _limit_to_sector_universe(df: pd.DataFrame) -> pd.DataFrame:
    if "broad_sector" not in df.columns:
        return df.copy()
    return df[df["broad_sector"].notna()].copy().reset_index(drop=True)


def _sector_relative_score(df: pd.DataFrame, column: str, direction: int) -> pd.Series:
    values = pd.to_numeric(df[column], errors="coerce")
    groups = df["broad_sector"] if "broad_sector" in df.columns else pd.Series("Universe", index=df.index)
    score = pd.Series(pd.NA, index=df.index, dtype="Float64")

    for _, index in groups.fillna("Unknown").groupby(groups.fillna("Unknown")).groups.items():
        group_values = values.loc[index]
        valid = group_values.dropna()
        if valid.empty:
            continue
        p10 = valid.quantile(0.10)
        p90 = valid.quantile(0.90)
        capped = group_values.clip(lower=p10, upper=p90)
        scaled = pd.Series(pd.NA, index=index, dtype="Float64")
        if p90 == p10:
            scaled.loc[valid.index] = 100.0
        else:
            scaled.loc[valid.index] = (capped.loc[valid.index] - p10) / (p90 - p10) * 100
        if direction < 0:
            scaled = 100 - scaled
        score.loc[index] = scaled

    return score.astype("Float64")


def _de_declining_yoy(df: pd.DataFrame) -> pd.Series:
    if "company_id" not in df.columns or "year" not in df.columns or "debt_to_equity" not in df.columns:
        return pd.Series(0, index=df.index)

    flags = pd.Series(0, index=df.index)
    annual = df[~df["year"].astype("string").str.strip().str.casefold().eq("ttm")].copy()
    annual["_year_sort_key"] = annual["year"].map(fiscal_year_sort_key)
    annual = annual.sort_values(["company_id", "_year_sort_key"])
    annual["_previous_de"] = annual.groupby("company_id")["debt_to_equity"].shift(1)
    current = pd.to_numeric(annual["debt_to_equity"], errors="coerce")
    previous = pd.to_numeric(annual["_previous_de"], errors="coerce")
    flags.loc[annual.index] = (current < previous).fillna(False).astype(int)
    return flags


def _fcf_cagr(df: pd.DataFrame, years: int = 5) -> pd.Series:
    if "company_id" not in df.columns or "year" not in df.columns or "free_cash_flow_cr" not in df.columns:
        return pd.Series(pd.NA, index=df.index, dtype="Float64")

    values = pd.Series(pd.NA, index=df.index, dtype="Float64")
    annual = df[~df["year"].astype("string").str.strip().str.casefold().eq("ttm")].copy()
    annual["_year_sort_key"] = annual["year"].map(fiscal_year_sort_key)
    for _, group in annual.sort_values("_year_sort_key").groupby("company_id", sort=False):
        fcf_values = pd.to_numeric(group["free_cash_flow_cr"], errors="coerce").tolist()
        group_indices = group.index.tolist()
        for position, index in enumerate(group_indices):
            values.loc[index] = calculate_cagr(fcf_values[position - years] if position >= years else None, fcf_values[position], years).value
    return values


def _sheet_name(preset_name: str) -> str:
    return preset_name.replace("_", " ").title()[:31]


def _colour_code_sheet(worksheet: Any, output: pd.DataFrame, thresholds: dict[str, float]) -> None:
    green_fill = PatternFill(fill_type="solid", fgColor="C6EFCE")
    red_fill = PatternFill(fill_type="solid", fgColor="FFC7CE")
    headers = {cell.value: cell.column for cell in worksheet[1]}

    for filter_name, threshold in thresholds.items():
        spec = _get_filter_spec(filter_name)
        column = next((alias for alias in spec.aliases if alias in output.columns), None)
        if column is None or column not in headers:
            continue
        excel_col = headers[column]
        values = pd.to_numeric(output[column], errors="coerce")
        for row_offset, value in enumerate(values, start=2):
            if pd.isna(value):
                continue
            if spec.direction == "min":
                passed = value >= threshold
            elif spec.direction == "max":
                passed = value <= threshold
            else:
                passed = value == threshold
            worksheet.cell(row=row_offset, column=excel_col).fill = green_fill if passed else red_fill


def _format_sheet(worksheet: Any) -> None:
    worksheet.freeze_panes = "A2"
    for column_cells in worksheet.columns:
        header = str(column_cells[0].value or "")
        max_length = max(len(str(cell.value or "")) for cell in column_cells[:100])
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, len(header) + 2), 28)


def sort_screener_output(df: pd.DataFrame) -> pd.DataFrame:
    """Sort screened companies by quality score, then stable identifiers."""
    sort_columns = ["composite_quality_score"]
    ascending = [False]
    for column in ("company_id", "year"):
        if column in df.columns:
            sort_columns.append(column)
            ascending.append(True)

    return df.sort_values(sort_columns, ascending=ascending, na_position="last").reset_index(drop=True)


def _normalise_thresholds(thresholds: dict[str, Any]) -> dict[str, float]:
    normalised = {}
    for key, value in thresholds.items():
        if value is None:
            continue
        normalised[str(key)] = float(value)
    return normalised


def _get_filter_spec(filter_name: str) -> FilterSpec:
    if filter_name not in FILTER_SPECS:
        supported = ", ".join(sorted(FILTER_SPECS))
        raise KeyError(f"Unsupported screener filter '{filter_name}'. Supported filters: {supported}")
    return FILTER_SPECS[filter_name]


def _resolve_column(df: pd.DataFrame, aliases: tuple[str, ...]) -> str:
    for alias in aliases:
        if alias in df.columns:
            return alias
    raise KeyError(f"Missing required screener column. Expected one of: {', '.join(aliases)}")


def _apply_financials_de_skip(df: pd.DataFrame, filter_mask: pd.Series) -> pd.Series:
    sector_column = "broad_sector" if "broad_sector" in df.columns else "sector" if "sector" in df.columns else None
    if sector_column is None:
        raise KeyError("D/E filter requires a broad_sector or sector column to skip Financials")

    is_financial = df[sector_column].astype("string").str.casefold().eq("financials")
    return filter_mask | is_financial.fillna(False)


def _apply_debt_free_icr_pass(df: pd.DataFrame, filter_mask: pd.Series) -> pd.Series:
    if "icr_label" not in df.columns:
        return filter_mask

    is_debt_free = df["icr_label"].astype("string").str.casefold().eq("debt free")
    return filter_mask | is_debt_free.fillna(False)
